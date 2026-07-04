from __future__ import annotations

import argparse
import json
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from PIL import Image


HEADERS = {
    "status": "检查结果",
    "images": "错误题目截图",
    "answers": "正确答案",
    "analysis": "错误分析",
}

BASE_HEADERS = ["题号", "URL", "备注1", "备注2"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Write Codex visual-review results and error images to Excel."
    )
    parser.add_argument("--excel", default="questions.xlsx")
    parser.add_argument("--results", default="_codex_review/results.json")
    parser.add_argument("--output", help="Output workbook; defaults to overwriting --excel")
    return parser.parse_args()


def ensure_headers(sheet) -> dict[str, int]:
    row_offset = 0
    first_url = sheet.cell(1, 2).value
    if isinstance(first_url, str) and first_url.strip().startswith("http"):
        sheet.insert_rows(1)
        for col, name in enumerate(BASE_HEADERS, start=1):
            sheet.cell(1, col).value = name
        row_offset = 1
    elif [sheet.cell(1, col).value for col in range(1, 5)] == BASE_HEADERS:
        row_offset = 1
    existing = {
        str(sheet.cell(1, col).value).strip(): col
        for col in range(1, sheet.max_column + 1)
        if sheet.cell(1, col).value is not None
    }
    for name in HEADERS.values():
        if name not in existing:
            col = sheet.max_column + 1
            sheet.cell(1, col).value = name
            existing[name] = col
    existing["_row_offset"] = row_offset
    return existing


def add_excel_images(sheet, row: int, col: int, paths: list[Path]) -> int:
    emu = 9525
    max_width = 520
    padding = 8
    y_offset = 0
    added = 0
    for path in paths:
        if not path.exists():
            continue
        with Image.open(path) as image:
            width, height = image.size
        scale = min(1.0, max_width / max(1, width))
        width = int(width * scale)
        height = int(height * scale)
        excel_image = ExcelImage(str(path))
        excel_image.width = width
        excel_image.height = height
        marker = AnchorMarker(col=col - 1, row=row - 1, rowOff=int(y_offset * emu))
        excel_image.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(width * emu, height * emu),
        )
        sheet.add_image(excel_image)
        y_offset += height + padding
        added += 1
    if y_offset:
        sheet.row_dimensions[row].height = max(24, y_offset * 0.75 + 8)
    return added


def load_results(path: Path) -> dict[int, dict[str, Any]]:
    raw = json.loads(path.read_text(encoding="utf-8-sig"))
    results: dict[int, dict[str, Any]] = {}
    for row, value in raw.items():
        item = dict(value)
        item["error_images"] = [
            Path(image_path) for image_path in item.get("error_images") or []
        ]
        results[int(row)] = item
    return results


def write_workbook(
    source: Path,
    destination: Path,
    results: dict[int, dict[str, Any]],
) -> None:
    workbook = load_workbook(source)
    sheet = workbook.active
    sheet._images = []
    headers = ensure_headers(sheet)
    row_offset = int(headers.pop("_row_offset", 0))
    status_col = headers[HEADERS["status"]]
    image_col = headers[HEADERS["images"]]
    answer_col = headers[HEADERS["answers"]]
    analysis_col = headers[HEADERS["analysis"]]

    for row in range(2, sheet.max_row + 1):
        for col in (status_col, image_col, answer_col, analysis_col):
            sheet.cell(row, col).value = None
        sheet.row_dimensions[row].height = 24

    for row, result in sorted(results.items()):
        target_row = row + row_offset
        sheet.cell(target_row, status_col).value = result.get("status", "")
        if result.get("correct_answers"):
            cell = sheet.cell(target_row, answer_col)
            cell.value = result["correct_answers"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if result.get("analysis"):
            cell = sheet.cell(target_row, analysis_col)
            cell.value = result["analysis"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        add_excel_images(
            sheet,
            target_row,
            image_col,
            result.get("error_images") or [],
        )

    sheet.column_dimensions[get_column_letter(status_col)].width = 38
    sheet.column_dimensions[get_column_letter(image_col)].width = 78
    sheet.column_dimensions[get_column_letter(answer_col)].width = 48
    sheet.column_dimensions[get_column_letter(analysis_col)].width = 72

    temp = destination.with_name(f".{destination.stem}.tmp{destination.suffix}")
    workbook.save(temp)
    workbook.close()
    os.replace(temp, destination)


def main() -> int:
    args = parse_args()
    workdir = Path.cwd()
    source = (workdir / args.excel).resolve()
    destination = (workdir / (args.output or args.excel)).resolve()
    results_path = (workdir / args.results).resolve()
    results = load_results(results_path)
    write_workbook(source, destination, results)
    print(f"已写回 {len(results)} 个 URL：{destination}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
