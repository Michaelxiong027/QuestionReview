from __future__ import annotations

import argparse
import concurrent.futures
import json
import re
import shutil
import threading
from dataclasses import asdict, dataclass
from datetime import datetime
from html import unescape
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qs, urlparse

import requests
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


PLATFORM_API = (
    "https://api-internal.tipaipai.com/maliang-inner-service/jf/page/final/info"
    "?page_id={page_id}&book_id={book_id}"
)
LOG_LOCK = threading.Lock()


@dataclass
class ExcelRow:
    row: int
    page_number: Any
    url: str
    page_id: str
    book_id: str


@dataclass
class QuestionRecord:
    sequence: str
    platform_answers: list[str]
    answer_image: str
    question_only_image: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Download platform payloads and render worksheet screenshots."
    )
    parser.add_argument("--excel", default="questions.xlsx")
    parser.add_argument("--authorization", default="authorization.txt")
    parser.add_argument("--screenshots-root", default="screenshots")
    parser.add_argument("--run-id", help="Output directory name under screenshots/")
    parser.add_argument("--url", action="append", help="Process an explicit URL; repeat as needed")
    parser.add_argument("--limit", type=int, help="Only process the first N URLs")
    parser.add_argument("--workers", type=int, default=4)
    return parser.parse_args()


def log(message: str) -> None:
    with LOG_LOCK:
        print(message, flush=True)


def parse_page_and_book(url: str) -> tuple[str, str]:
    parsed = urlparse(url.strip())
    query = parse_qs(parsed.query)
    if parsed.fragment and "?" in parsed.fragment:
        query.update(parse_qs(parsed.fragment.split("?", 1)[1]))
    page_id = (query.get("page_id") or [None])[0]
    book_id = (query.get("book_id") or [None])[0]
    if not page_id or not book_id:
        raise ValueError(f"URL missing page_id or book_id: {url}")
    return str(page_id), str(book_id)


def find_header(sheet, names: set[str]) -> int | None:
    normalized = {name.lower() for name in names}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(1, col).value
        if isinstance(value, str) and value.strip().lower() in normalized:
            return col
    return None


def read_excel_rows(path: Path, explicit_urls: list[str] | None) -> list[ExcelRow]:
    if explicit_urls:
        return [
            ExcelRow(index + 2, None, url, *parse_page_and_book(url))
            for index, url in enumerate(explicit_urls)
        ]

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    url_col = find_header(sheet, {"url", "urls", "link"})
    page_col = find_header(sheet, {"页码", "题号", "page", "page_number"})
    if url_col is None:
        for col in range(1, sheet.max_column + 1):
            for row in range(1, min(sheet.max_row, 20) + 1):
                value = sheet.cell(row, col).value
                if isinstance(value, str) and value.strip().startswith("http"):
                    url_col = col
                    if page_col is None and col > 1:
                        page_col = col - 1
                    break
            if url_col is not None:
                break
    if url_col is None:
        workbook.close()
        raise ValueError("No URL column found in workbook")

    rows: list[ExcelRow] = []
    start_row = 2 if find_header(sheet, {"url", "urls", "link"}) else 1
    for row in range(start_row, sheet.max_row + 1):
        value = sheet.cell(row, url_col).value
        if not isinstance(value, str) or not value.strip().startswith("http"):
            continue
        page_id, book_id = parse_page_and_book(value)
        rows.append(
            ExcelRow(
                row=row,
                page_number=sheet.cell(row, page_col).value if page_col else None,
                url=value.strip(),
                page_id=page_id,
                book_id=book_id,
            )
        )
    workbook.close()
    return rows


def read_authorization(path: Path) -> str:
    token = path.read_text(encoding="utf-8-sig").strip()
    if not token:
        raise ValueError(f"Authorization file is empty: {path}")
    return token


def fetch_platform_payload(page_id: str, book_id: str, authorization: str) -> dict:
    response = requests.get(
        PLATFORM_API.format(page_id=page_id, book_id=book_id),
        headers={
            "Authorization": authorization,
            "Accept": "application/json, text/plain, */*",
            "User-Agent": "Mozilla/5.0",
        },
        timeout=45,
    )
    response.raise_for_status()
    payload = response.json()
    if payload.get("error_code") != 0:
        raise RuntimeError(payload.get("error_msg") or "Platform API returned an error")
    return payload


def download_image(url: str) -> Image.Image:
    response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=90)
    response.raise_for_status()
    return Image.open(BytesIO(response.content)).convert("RGB")


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for path in (
        Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("C:/Windows/Fonts/arial.ttf"),
    ):
        if path.exists():
            return ImageFont.truetype(str(path), size)
    return ImageFont.load_default()


def plain_text(value: Any) -> str:
    if value is None:
        return ""
    return unescape(re.sub(r"<[^>]+>", "", str(value))).strip()


def polygon_to_bbox(
    polygon: Iterable[dict[str, Any]], width: int, height: int
) -> tuple[int, int, int, int]:
    points = list(polygon)
    xs = [float(point["x"]) * width for point in points]
    ys = [float(point["y"]) * height for point in points]
    return (
        max(0, int(round(min(xs)))),
        max(0, int(round(min(ys)))),
        min(width, int(round(max(xs)))),
        min(height, int(round(max(ys)))),
    )


def union_bbox(boxes: list[tuple[int, int, int, int]]) -> tuple[int, int, int, int]:
    return (
        min(box[0] for box in boxes),
        min(box[1] for box in boxes),
        max(box[2] for box in boxes),
        max(box[3] for box in boxes),
    )


def get_answer_items(container: dict[str, Any]) -> list[dict[str, Any]]:
    answers: list[dict[str, Any]] = []
    for key in ("answerInfo", "answerInfo2"):
        value = container.get(key)
        if isinstance(value, list):
            answers.extend(item for item in value if isinstance(item, dict))
    independent = (container.get("answerIndependent") or {}).get("iAnswerInfo")
    if isinstance(independent, list):
        answers.extend(item for item in independent if isinstance(item, dict))
    elif isinstance(independent, dict):
        answers.append(independent)
    return answers


def answer_bbox(
    answer: dict[str, Any], width: int, height: int
) -> tuple[int, int, int, int] | None:
    coordinates = (answer.get("area") or {}).get("coordinates")
    return polygon_to_bbox(coordinates, width, height) if coordinates else None


def fit_text(
    draw: ImageDraw.ImageDraw,
    text: str,
    max_width: int,
    preferred_size: int,
    minimum_size: int = 7,
    max_height: int | None = None,
) -> tuple[str, ImageFont.FreeTypeFont | ImageFont.ImageFont, int]:
    value = plain_text(text)
    for size in range(preferred_size, minimum_size - 1, -1):
        font = load_font(size)
        box = draw.textbbox((0, 0), value, font=font)
        width = box[2] - box[0]
        height = box[3] - box[1]
        if width <= max_width and (max_height is None or height <= max_height):
            return value, font, width
    font = load_font(minimum_size)
    clipped = value
    while clipped:
        candidate = clipped + ("..." if clipped != value else "")
        box = draw.textbbox((0, 0), candidate, font=font)
        width = box[2] - box[0]
        height = box[3] - box[1]
        if width <= max_width and (max_height is None or height <= max_height):
            return candidate, font, width
        clipped = clipped[:-1]
    return "", font, 0


def draw_answer_content(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    answer_text: str,
    label_text: str,
    preferred_size: int,
) -> None:
    left, top, right, bottom = box
    inner_width = max(1, right - left - 8)
    inner_height = max(1, bottom - top - 4)
    box_size = max(10, min(24, int(round(inner_height * 0.78))))
    preferred = max(preferred_size, box_size)
    minimum = max(7, min(12, int(round(inner_height * 0.45))))
    max_text_height = max(7, inner_height)
    label, label_font, label_width = fit_text(
        draw,
        label_text,
        max(40, int(inner_width * 0.46)),
        preferred,
        minimum,
        max_text_height,
    )
    answer, answer_font, _ = fit_text(
        draw,
        answer_text,
        max(1, inner_width - label_width - 12),
        preferred,
        minimum,
        max_text_height,
    )
    label_box = draw.textbbox((0, 0), label, font=label_font)
    label_height = label_box[3] - label_box[1]
    label_x = max(left + 4, right - label_width - 4)
    label_y = top + 2
    draw.rectangle(
        (
            label_x - 2,
            label_y - 1,
            right - 2,
            min(bottom - 1, label_y + label_height + 2),
        ),
        fill="#EDF5FF",
    )
    draw.text((label_x, label_y), label, fill="#880015", font=label_font)
    if answer:
        draw.text((left + 4, top + 2), answer, fill="#0078D4", font=answer_font)

def safe_sequence(value: Any, fallback: int) -> str:
    text = str(value if value not in (None, "") else fallback).strip()
    return re.sub(r"[^0-9A-Za-z_.-]+", "_", text) or str(fallback)


def render_page(payload: dict, page_dir: Path) -> tuple[Path, list[QuestionRecord]]:
    data = payload.get("data") or {}
    label_items = data.get("label_containers") or []
    image_url = data.get("file_path")
    if not image_url and label_items:
        first = (label_items[0].get("labelContainer") or {}).get("question") or [{}]
        image_url = first[0].get("filePath")
    if not image_url:
        raise RuntimeError("Platform response has no source page image")

    page_dir.mkdir(parents=True, exist_ok=True)
    answer_dir = page_dir / "answer"
    question_only_dir = page_dir / "question_only"
    answer_dir.mkdir(exist_ok=True)
    question_only_dir.mkdir(exist_ok=True)

    source = download_image(str(image_url))
    source_path = page_dir / "source_page.jpg"
    source.save(source_path, quality=95)
    page_w, page_h = source.size
    preferred_size = max(8, min(14, round(page_w / 220)))
    number_font = load_font(max(13, preferred_size + 3))
    records: list[QuestionRecord] = []

    for index, item in enumerate(label_items, start=1):
        container = item.get("labelContainer") or {}
        sequence = safe_sequence(container.get("sequence"), index)
        answer_items: list[tuple[dict[str, Any], tuple[int, int, int, int], str]] = []
        for answer in get_answer_items(container):
            answer_text = plain_text(answer.get("answerText"))
            if not answer_text.strip():
                continue
            box = answer_bbox(answer, page_w, page_h)
            if box is None:
                continue
            answer_items.append((answer, box, answer_text))
        if not answer_items:
            continue

        questions = container.get("question") or []
        question_boxes = [
            polygon_to_bbox(question["coordinates"], page_w, page_h)
            for question in questions
            if question.get("coordinates")
        ]
        if not question_boxes:
            continue

        crop_box = union_bbox(question_boxes)
        question_only_path = question_only_dir / f"{index:03d}_question_{sequence}.png"
        source.crop(crop_box).save(question_only_path)

        answer_image = source.crop(crop_box)
        draw = ImageDraw.Draw(answer_image)
        offset_x, offset_y = crop_box[0], crop_box[1]
        for question_box in question_boxes:
            local = (
                question_box[0] - offset_x,
                question_box[1] - offset_y,
                question_box[2] - offset_x,
                question_box[3] - offset_y,
            )
            draw.rectangle(local, outline="#4D9CFF", width=3)

        platform_answers: list[str] = []
        for answer_index, (_, box, answer_text) in enumerate(answer_items, start=1):
            platform_answers.append(answer_text)
            local = (
                box[0] - offset_x,
                box[1] - offset_y,
                box[2] - offset_x,
                box[3] - offset_y,
            )
            draw.rectangle(local, outline="#FFD400", width=3)
            draw_answer_content(
                draw,
                local,
                answer_text,
                f"{sequence}:\u7b54{answer_index}",
                preferred_size,
            )

        number_box = draw.textbbox((0, 0), sequence, font=number_font)
        number_w = number_box[2] - number_box[0]
        number_h = number_box[3] - number_box[1]
        number_x = max(0, answer_image.width - number_w - 3)
        draw.rectangle(
            (number_x - 2, 0, answer_image.width - 1, number_h + 4),
            fill="#EDF5FF",
        )
        draw.text((number_x, 1), sequence, fill="#00A046", font=number_font)

        answer_path = answer_dir / f"{index:03d}_question_{sequence}.png"
        answer_image.save(answer_path)
        records.append(
            QuestionRecord(
                sequence=sequence,
                platform_answers=platform_answers,
                answer_image=str(answer_path),
                question_only_image=str(question_only_path),
            )
        )

    return source_path, records


def process_url(
    position: int,
    total: int,
    row: ExcelRow,
    authorization: str,
    run_dir: Path,
) -> dict[str, Any]:
    page_dir = run_dir / (
        f"url_{position:04d}_row={row.row}_"
        f"page_id={row.page_id}&book_id={row.book_id}"
    )
    try:
        if page_dir.exists():
            shutil.rmtree(page_dir)
        page_dir.mkdir(parents=True)

        payload = fetch_platform_payload(row.page_id, row.book_id, authorization)
        (page_dir / "platform_response.json").write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        source_page, questions = render_page(payload, page_dir)
        item = {
            **asdict(row),
            "position": position,
            "page_dir": str(page_dir),
            "source_page": str(source_page),
            "question_count": len(questions),
            "questions": [asdict(question) for question in questions],
            "status": "downloaded",
        }
        (page_dir / "page_manifest.json").write_text(
            json.dumps(item, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        log(f"[{position}/{total}] downloaded: page_id={row.page_id}, questions={len(questions)}")
        return item
    except Exception as exc:
        page_dir.mkdir(parents=True, exist_ok=True)
        (page_dir / "error.txt").write_text(str(exc), encoding="utf-8")
        log(f"[{position}/{total}] failed: {exc}")
        return {
            **asdict(row),
            "position": position,
            "page_dir": str(page_dir),
            "status": "error",
            "error": str(exc),
        }


def main() -> int:
    args = parse_args()
    workdir = Path.cwd()
    excel_path = (workdir / args.excel).resolve()
    authorization_path = (workdir / args.authorization).resolve()
    screenshots_root = (workdir / args.screenshots_root).resolve()
    run_id = args.run_id or datetime.now().strftime("run_%Y%m%d_%H%M%S")
    run_dir = screenshots_root / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    rows = read_excel_rows(excel_path, args.url)
    if args.limit is not None:
        rows = rows[: max(0, args.limit)]
    if not rows:
        log("No valid URL found.")
        return 2

    authorization = read_authorization(authorization_path)
    workers = max(1, int(args.workers))
    manifest: list[dict[str, Any]] = []
    log(f"Run directory: {run_dir}")
    log(f"URLs: {len(rows)}; workers: {workers}")

    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(process_url, i, len(rows), row, authorization, run_dir): i
            for i, row in enumerate(rows, start=1)
        }
        for future in concurrent.futures.as_completed(futures):
            manifest.append(future.result())

    manifest.sort(key=lambda item: int(item["position"]))
    (run_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    failed = sum(item["status"] == "error" for item in manifest)
    log(f"Finished: {len(manifest) - failed} succeeded, {failed} failed.")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())

